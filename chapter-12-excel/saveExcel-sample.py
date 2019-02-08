#! python3
# -*- coding:utf8 -*-

import sys,os,time
import openpyxl as excel

wb = excel.Workbook()
sheet = wb.get_active_sheet()
print(wb.get_sheet_names())

sheet.title = 'Span sPan spAn spaN'
print(wb.get_sheet_names())

wb.create_sheet(index= 0,title='First sheet')
print(wb.get_sheet_names())

wb.create_sheet(title='Last sheet')
print(wb.get_sheet_names())

wb.remove_sheet(wb.get_sheet_by_name('First sheet'))
print(wb.get_sheet_names())

wb.remove_sheet(wb.get_sheet_by_name('Last sheet'))
print(wb.get_sheet_names())

wb.create_sheet()
print(wb.get_sheet_names())
sheet = wb.get_active_sheet()
print(sheet.title)

sheet = wb.get_sheet_by_name('Sheet')
print(sheet.title)

print(sheet['A1'].value)
sheet['A1'] = 'Hello openpyxl/excel'
print(sheet['A1'].value)

#wb.save('example-sample.xlsx')

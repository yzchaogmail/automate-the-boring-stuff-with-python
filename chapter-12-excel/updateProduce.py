#! python3
# -*- coding:utf-8 -*-

import openpyxl,os,sys,logging

#logging.disable(logging.CRITICAL)
logging.basicConfig(filename = 'debuglog.txt',level =
           logging.DEBUG,format='%(asctime)s-%(levelname)s-%(message)s')

PRODUCT_UPDATES = {'Garlic':3.07,
                 'Celery':1.19,
                 'Lemon':1.27}

wb  = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_active_sheet()

logging.debug('%s ' %sheet.title )
logging.debug('%d ' %sheet.max_row )

for i in range(1,sheet.max_row+1):
    productName = sheet.cell(column = 1,row = i).value

    if(productName in PRODUCT_UPDATES.keys()):
        logging.debug(productName)
        logging.debug('Before = %f' %(sheet.cell(column = 2,row = i).value))

        sheet['B'+str(i)] = PRODUCT_UPDATES[productName]

        logging.debug('After = %f' %(sheet['B'+str(i)].value))

wb.save('updatedProductSals.xlsx')

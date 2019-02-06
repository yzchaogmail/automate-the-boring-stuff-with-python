#! python3
# -*- codeing=utf-8 -*-
import sys,os
import openpyxl as excel
from openpyxl.utils import get_column_letter,column_index_from_string
#os.getcwd()
#print(os.getcwd())
#path = os.chdir(os.getcwd())
print(os.getcwd())

try:
    workbook = excel.load_workbook('example.xlsx')
    print('Success load workbook "example.xlsx"')
    print(workbook.get_sheet_names())
    #print(os.getcwd())
    #
    #获取表工作簿
    #
    '''
    try:
        sheet = workbook.get_sheet_by_name('Sheet3')
        # print(type(sheet))
        print('get_sheet_by_name Sheet-title:%s' %sheet.title)
    except:
        print('Failed to get sheet by name.')
    '''
    sheet = workbook.get_active_sheet()
    #
    #表中取得单元格
    #
    print(sheet['A1'].value)
    print(sheet['B1'].value)
    # highestColumn = sheet.get_highest_column()
    # print('Highest column:%f' %get_highest_column)
    print('Max-Row=%d' %sheet.max_row)
    print('Max-Column=%d' %sheet.max_column)

    cell = sheet['B1']
    print(cell.row)
    print(cell.column)
    print(cell.coordinate)
    for i in range(1,8,2):
        cell = sheet.cell(row = i,column=2)
        print(i,cell.row,cell.column,cell.value)
    #
    # 列字母与数字之间的转换
    #
    print(get_column_letter(27))
    print(column_index_from_string('BB'))

    #
    # 遍历工作簿中的单元格
    #
    wb = excel.load_workbook('example.xlsx')
    sheet = wb.get_active_sheet()
    for rawCells in sheet['A1':'C3']:
        #print(rawCells.row)
        print(tuple(rawCells))
        for cellObj in rawCells:
            print(cellObj.coordinate,cellObj.value)
        print('---End of row---')
    #
    #list(sheet.columns/sheet.rows)遍历sheet工作簿
    #
    wb = excel.load_workbook('example.xlsx')
    sheet = wb.get_active_sheet()
    print('------>>>>>sheet.columns[1] way<<<<<<-----------')
    for cellObj in list(sheet.columns)[0]:
        print(cellObj.coordinate,cellObj.value)
except:
    print('Failed to load workbook "example.xlsx"')
